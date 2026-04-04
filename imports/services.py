from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from catalog.models import Legend, Set
from tournaments.models import (
    Match,
    Round,
    RoundStatus,
    Tournament,
    TournamentEntry,
    TournamentStatus,
)
from users.models import Player


REQUIRED_SHEETS = (
    "tournament",
    "players",
    "entries",
    "rounds",
    "matches",
)

REQUIRED_COLUMNS = {
    "tournament": (
        "name",
        "slug",
        "date",
        "status",
        "set_slug",
        "external_id",
        "notes",
    ),
    "players": (
        "player_external_id",
        "display_name",
        "slug",
        "name",
        "avatar",
    ),
    "entries": (
        "player_external_id",
        "legend_slug",
        "wins",
        "losses",
        "draws",
        "points",
        "final_position",
    ),
    "rounds": (
        "number",
        "status",
        "external_id",
    ),
    "matches": (
        "round_number",
        "table_number",
        "player1_external_id",
        "player2_external_id",
        "player1_wins",
        "player2_wins",
        "external_id",
    ),
}


class TournamentImportError(Exception):
    pass


@dataclass
class WorkbookValidationResult:
    sheet_row_counts: dict[str, int]
    parsed_data: dict[str, list[dict[str, Any]]]


@dataclass
class TournamentImportResult:
    tournament: Tournament
    created_counts: dict[str, int]
    updated_counts: dict[str, int]


class TournamentWorkbookValidator:
    def __init__(self, uploaded_file):
        self.uploaded_file = uploaded_file
        self.workbook = None

    def run(self) -> WorkbookValidationResult:
        self.workbook = self._load_workbook()
        self._validate_required_sheets()

        parsed_data: dict[str, list[dict[str, Any]]] = {}
        sheet_row_counts: dict[str, int] = {}

        for sheet_name in REQUIRED_SHEETS:
            worksheet = self.workbook[sheet_name]
            rows = self._parse_sheet(worksheet, sheet_name)
            parsed_data[sheet_name] = rows
            sheet_row_counts[sheet_name] = len(rows)

        self._validate_minimum_content(parsed_data)

        return WorkbookValidationResult(
            sheet_row_counts=sheet_row_counts,
            parsed_data=parsed_data,
        )

    def _load_workbook(self):
        try:
            self.uploaded_file.seek(0)
            return load_workbook(self.uploaded_file, data_only=True)
        except Exception as exc:
            raise TournamentImportError(
                "No se ha podido abrir el archivo Excel. Asegúrate de subir un .xlsx válido."
            ) from exc

    def _validate_required_sheets(self):
        existing_sheets = set(self.workbook.sheetnames)
        missing_sheets = [name for name in REQUIRED_SHEETS if name not in existing_sheets]

        if missing_sheets:
            missing = ", ".join(missing_sheets)
            raise TournamentImportError(
                f"Faltan hojas obligatorias en el Excel: {missing}."
            )

    def _parse_sheet(self, worksheet, sheet_name: str) -> list[dict[str, Any]]:
        header = self._read_header(worksheet, sheet_name)
        required_columns = REQUIRED_COLUMNS[sheet_name]
        self._validate_required_columns(sheet_name, header, required_columns)

        rows: list[dict[str, Any]] = []

        for excel_row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            normalized_values = [self._normalize_cell_value(value) for value in row_values]

            if self._is_empty_row(normalized_values):
                continue

            row_dict = {}
            for index, column_name in enumerate(header):
                row_dict[column_name] = normalized_values[index] if index < len(normalized_values) else ""

            row_dict["__row_number__"] = excel_row_number
            rows.append(row_dict)

        return rows

    def _read_header(self, worksheet, sheet_name: str) -> list[str]:
        header_values = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )

        if not header_values:
            raise TournamentImportError(
                f"La hoja '{sheet_name}' no tiene fila de cabecera."
            )

        header = [self._normalize_header_value(value) for value in header_values]

        if not any(header):
            raise TournamentImportError(
                f"La hoja '{sheet_name}' no tiene una cabecera válida."
            )

        return header

    def _validate_required_columns(
        self,
        sheet_name: str,
        header: list[str],
        required_columns: tuple[str, ...],
    ):
        missing_columns = [column for column in required_columns if column not in header]

        if missing_columns:
            missing = ", ".join(missing_columns)
            raise TournamentImportError(
                f"Faltan columnas obligatorias en la hoja '{sheet_name}': {missing}."
            )

    def _validate_minimum_content(self, parsed_data: dict[str, list[dict[str, Any]]]):
        tournament_rows = parsed_data["tournament"]

        if len(tournament_rows) != 1:
            raise TournamentImportError(
                "La hoja 'tournament' debe tener exactamente 1 fila de datos."
            )

    @staticmethod
    def _normalize_header_value(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    def _normalize_cell_value(self, value: Any) -> Any:
        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.date().isoformat()

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, str):
            return value.strip()

        return value

    @staticmethod
    def _is_empty_row(values: list[Any]) -> bool:
        for value in values:
            if value not in ("", None):
                return False
        return True


class TournamentWorkbookImporter:
    def __init__(self, validation_result: WorkbookValidationResult):
        self.data = validation_result.parsed_data

        self.created_counts = {
            "players": 0,
            "entries": 0,
            "rounds": 0,
            "matches": 0,
            "tournaments": 0,
        }
        self.updated_counts = {
            "players": 0,
            "entries": 0,
            "rounds": 0,
            "matches": 0,
            "tournaments": 0,
        }

        self.player_by_external_id: dict[str, Player] = {}
        self.round_by_number: dict[int, Round] = {}
        self.entry_by_player_external_id: dict[str, TournamentEntry] = {}

    def run(self) -> TournamentImportResult:
        self._validate_semantics()

        with transaction.atomic():
            tournament = self._import_tournament()
            self._import_players()
            self._import_entries(tournament)
            self._import_rounds(tournament)
            self._import_matches()

        return TournamentImportResult(
            tournament=tournament,
            created_counts=self.created_counts,
            updated_counts=self.updated_counts,
        )

    def _validate_semantics(self):
        self._validate_tournament_row(self.data["tournament"][0])
        self._validate_players_sheet(self.data["players"])
        self._validate_entries_sheet(self.data["entries"])
        self._validate_rounds_sheet(self.data["rounds"])
        self._validate_matches_sheet(self.data["matches"])

    def _validate_tournament_row(self, row: dict[str, Any]):
        self._require_value(row, "name", "tournament")
        self._require_value(row, "slug", "tournament")
        self._require_value(row, "date", "tournament")
        self._require_value(row, "status", "tournament")
        self._require_value(row, "set_slug", "tournament")

        status = str(row["status"]).strip().lower()
        if status not in {TournamentStatus.DRAFT, TournamentStatus.FINISHED}:
            raise TournamentImportError(
                f"Hoja 'tournament', fila {row['__row_number__']}: "
                f"status inválido '{row['status']}'. Usa 'draft' o 'finished'."
            )

        set_slug = str(row["set_slug"]).strip()
        if not Set.objects.filter(slug=set_slug).exists():
            raise TournamentImportError(
                f"Hoja 'tournament', fila {row['__row_number__']}: "
                f"no existe ningún set con slug '{set_slug}'."
            )

    def _validate_players_sheet(self, rows: list[dict[str, Any]]):
        seen_external_ids = set()
        seen_slugs = set()

        for row in rows:
            self._require_value(row, "player_external_id", "players")
            self._require_value(row, "display_name", "players")
            self._require_value(row, "slug", "players")

            external_id = str(row["player_external_id"]).strip()
            slug = str(row["slug"]).strip()

            if external_id in seen_external_ids:
                raise TournamentImportError(
                    f"Hoja 'players', fila {row['__row_number__']}: "
                    f"player_external_id duplicado '{external_id}'."
                )
            if slug in seen_slugs:
                raise TournamentImportError(
                    f"Hoja 'players', fila {row['__row_number__']}: "
                    f"slug duplicado '{slug}'."
                )

            seen_external_ids.add(external_id)
            seen_slugs.add(slug)

    def _validate_entries_sheet(self, rows: list[dict[str, Any]]):
        player_ids = {str(row["player_external_id"]).strip() for row in self.data["players"]}
        seen_entry_players = set()

        for row in rows:
            self._require_value(row, "player_external_id", "entries")

            player_external_id = str(row["player_external_id"]).strip()
            legend_slug = str(row.get("legend_slug", "")).strip()

            if player_external_id not in player_ids:
                raise TournamentImportError(
                    f"Hoja 'entries', fila {row['__row_number__']}: "
                    f"player_external_id '{player_external_id}' no existe en 'players'."
                )

            if player_external_id in seen_entry_players:
                raise TournamentImportError(
                    f"Hoja 'entries', fila {row['__row_number__']}: "
                    f"jugador duplicado en entries ('{player_external_id}')."
                )

            if legend_slug and not Legend.objects.filter(slug=legend_slug).exists():
                raise TournamentImportError(
                    f"Hoja 'entries', fila {row['__row_number__']}: "
                    f"no existe ninguna leyenda con slug '{legend_slug}'."
                )

            self._parse_non_negative_int(row, "wins", "entries")
            self._parse_non_negative_int(row, "losses", "entries")
            self._parse_non_negative_int(row, "draws", "entries")
            self._parse_non_negative_int(row, "points", "entries")

            if row.get("final_position", "") != "":
                self._parse_positive_int(row, "final_position", "entries")

            seen_entry_players.add(player_external_id)

    def _validate_rounds_sheet(self, rows: list[dict[str, Any]]):
        seen_numbers = set()

        for row in rows:
            number = self._parse_positive_int(row, "number", "rounds")
            self._require_value(row, "status", "rounds")

            status = str(row["status"]).strip().lower()
            if status not in {"draft", "pending", "finished"}:
                raise TournamentImportError(
                    f"Hoja 'rounds', fila {row['__row_number__']}: "
                    f"status inválido '{row['status']}'. Usa 'draft', 'pending' o 'finished'."
                )

            if number in seen_numbers:
                raise TournamentImportError(
                    f"Hoja 'rounds', fila {row['__row_number__']}: "
                    f"número de ronda duplicado '{number}'."
                )

            seen_numbers.add(number)

    def _validate_matches_sheet(self, rows: list[dict[str, Any]]):
        round_numbers = {
            self._parse_positive_int(row, "number", "rounds")
            for row in self.data["rounds"]
        }
        entry_player_ids = {
            str(row["player_external_id"]).strip()
            for row in self.data["entries"]
        }

        for row in rows:
            round_number = self._parse_positive_int(row, "round_number", "matches")
            self._require_value(row, "player1_external_id", "matches")

            player1_external_id = str(row["player1_external_id"]).strip()
            player2_external_id = str(row.get("player2_external_id", "")).strip()
            is_bye = player2_external_id == ""

            if round_number not in round_numbers:
                raise TournamentImportError(
                    f"Hoja 'matches', fila {row['__row_number__']}: "
                    f"round_number '{round_number}' no existe en 'rounds'."
                )

            if player1_external_id not in entry_player_ids:
                raise TournamentImportError(
                    f"Hoja 'matches', fila {row['__row_number__']}: "
                    f"player1_external_id '{player1_external_id}' no existe en 'entries'."
                )

            if not is_bye and player2_external_id not in entry_player_ids:
                raise TournamentImportError(
                    f"Hoja 'matches', fila {row['__row_number__']}: "
                    f"player2_external_id '{player2_external_id}' no existe en 'entries'."
                )

            if not is_bye and player1_external_id == player2_external_id:
                raise TournamentImportError(
                    f"Hoja 'matches', fila {row['__row_number__']}: "
                    "un jugador no puede enfrentarse a sí mismo."
                )

            self._parse_non_negative_int(row, "player1_wins", "matches")
            self._parse_non_negative_int(row, "player2_wins", "matches")

            if is_bye:
                player1_wins = self._parse_non_negative_int(row, "player1_wins", "matches")
                player2_wins = self._parse_non_negative_int(row, "player2_wins", "matches")

                if player1_wins < player2_wins:
                    raise TournamentImportError(
                        f"Hoja 'matches', fila {row['__row_number__']}: "
                        "en un BYE, player1_wins no puede ser menor que player2_wins."
                    )

            if row.get("table_number", "") != "":
                self._parse_positive_int(row, "table_number", "matches")

    def _import_tournament(self) -> Tournament:
        row = self.data["tournament"][0]
        external_id = self._as_str(row.get("external_id"))
        slug = self._as_str(row["slug"])

        tournament = None

        if external_id:
            tournament = Tournament.objects.filter(external_id=external_id).first()

        if tournament is None:
            tournament = Tournament.objects.filter(slug=slug).first()

        tournament_set = Set.objects.get(slug=self._as_str(row["set_slug"]))

        defaults = {
            "name": self._as_str(row["name"]),
            "slug": slug,
            "date": self._parse_date_value(row["date"], "tournament", row["__row_number__"]),
            "status": self._normalize_tournament_status(self._as_str(row["status"])),
            "set": tournament_set,
            "notes": self._as_str(row.get("notes")),
            "external_id": external_id or None,
        }

        if tournament is None:
            tournament = Tournament.objects.create(**defaults)
            self.created_counts["tournaments"] += 1
            return tournament

        for field, value in defaults.items():
            setattr(tournament, field, value)
        tournament.save()
        self.updated_counts["tournaments"] += 1
        return tournament

    def _import_players(self):
        for row in self.data["players"]:
            external_id = self._as_str(row["player_external_id"])
            slug = self._as_str(row["slug"])

            player = Player.objects.filter(external_id=external_id).first()
            if player is None:
                player = Player.objects.filter(slug=slug).first()

            defaults = {
                "display_name": self._as_str(row["display_name"]),
                "slug": slug,
                "external_id": external_id,
                "is_active": True,
            }

            if player is None:
                player = Player.objects.create(**defaults)
                self.created_counts["players"] += 1
            else:
                for field, value in defaults.items():
                    setattr(player, field, value)
                player.save()
                self.updated_counts["players"] += 1

            self.player_by_external_id[external_id] = player

    def _import_entries(self, tournament: Tournament):
        for row in self.data["entries"]:
            player_external_id = self._as_str(row["player_external_id"])
            player = self.player_by_external_id[player_external_id]

            legend_slug = self._as_str(row.get("legend_slug"))
            legend = Legend.objects.get(slug=legend_slug) if legend_slug else None

            entry = TournamentEntry.objects.filter(
                tournament=tournament,
                player=player,
            ).first()

            defaults = {
                "legend": legend,
                "wins": self._parse_non_negative_int(row, "wins", "entries"),
                "losses": self._parse_non_negative_int(row, "losses", "entries"),
                "draws": self._parse_non_negative_int(row, "draws", "entries"),
                "points": self._parse_non_negative_int(row, "points", "entries"),
                "final_position": (
                    self._parse_positive_int(row, "final_position", "entries")
                    if row.get("final_position", "") != ""
                    else None
                ),
                "external_id": None,
            }

            if entry is None:
                entry = TournamentEntry.objects.create(
                    tournament=tournament,
                    player=player,
                    **defaults,
                )
                self.created_counts["entries"] += 1
            else:
                for field, value in defaults.items():
                    setattr(entry, field, value)
                entry.save()
                self.updated_counts["entries"] += 1

            self.entry_by_player_external_id[player_external_id] = entry

    def _import_rounds(self, tournament: Tournament):
        for row in self.data["rounds"]:
            number = self._parse_positive_int(row, "number", "rounds")

            round_obj = Round.objects.filter(
                tournament=tournament,
                number=number,
            ).first()

            defaults = {
                "status": self._normalize_round_status(self._as_str(row["status"])),
                "external_id": self._as_str(row.get("external_id")) or None,
            }

            if round_obj is None:
                round_obj = Round.objects.create(
                    tournament=tournament,
                    number=number,
                    **defaults,
                )
                self.created_counts["rounds"] += 1
            else:
                for field, value in defaults.items():
                    setattr(round_obj, field, value)
                round_obj.save()
                self.updated_counts["rounds"] += 1

            self.round_by_number[number] = round_obj

    def _import_matches(self):
        for row in self.data["matches"]:
            external_id = self._as_str(row.get("external_id"))
            round_number = self._parse_positive_int(row, "round_number", "matches")
            round_obj = self.round_by_number[round_number]

            match = None
            if external_id:
                match = Match.objects.filter(external_id=external_id).first()

            player1_entry = self.entry_by_player_external_id[
                self._as_str(row["player1_external_id"])
            ]
            player2_external_id = self._as_str(row.get("player2_external_id"))
            is_bye = player2_external_id == ""

            defaults = {
                "round": round_obj,
                "table_number": (
                    self._parse_positive_int(row, "table_number", "matches")
                    if row.get("table_number", "") != ""
                    else None
                ),
                "player1_entry": player1_entry,
                "player2_entry": (
                    None
                    if is_bye
                    else self.entry_by_player_external_id[player2_external_id]
                ),
                "is_bye": is_bye,
                "player1_wins": 2 if is_bye else self._parse_non_negative_int(row, "player1_wins", "matches"),
                "player2_wins": 0 if is_bye else self._parse_non_negative_int(row, "player2_wins", "matches"),
                "external_id": external_id or None,
            }

            if match is None:
                match = Match(**defaults)
                match.full_clean()
                match.save()
                self.created_counts["matches"] += 1
            else:
                for field, value in defaults.items():
                    setattr(match, field, value)
                match.full_clean()
                match.save()
                self.updated_counts["matches"] += 1

    def _require_value(self, row: dict[str, Any], field_name: str, sheet_name: str):
        value = row.get(field_name, "")
        if self._as_str(value) == "":
            raise TournamentImportError(
                f"Hoja '{sheet_name}', fila {row['__row_number__']}: "
                f"el campo '{field_name}' es obligatorio."
            )

    def _parse_non_negative_int(self, row: dict[str, Any], field_name: str, sheet_name: str) -> int:
        value = row.get(field_name, "")

        try:
            parsed = int(value)
        except (TypeError, ValueError):
            raise TournamentImportError(
                f"Hoja '{sheet_name}', fila {row['__row_number__']}: "
                f"el campo '{field_name}' debe ser un entero."
            )

        if parsed < 0:
            raise TournamentImportError(
                f"Hoja '{sheet_name}', fila {row['__row_number__']}: "
                f"el campo '{field_name}' no puede ser negativo."
            )

        return parsed

    def _parse_positive_int(self, row: dict[str, Any], field_name: str, sheet_name: str) -> int:
        parsed = self._parse_non_negative_int(row, field_name, sheet_name)

        if parsed <= 0:
            raise TournamentImportError(
                f"Hoja '{sheet_name}', fila {row['__row_number__']}: "
                f"el campo '{field_name}' debe ser mayor que 0."
            )

        return parsed

    def _parse_date_value(self, value: Any, sheet_name: str, row_number: int):
        if isinstance(value, date):
            return value

        raw = self._as_str(value)
        try:
            return date.fromisoformat(raw)
        except ValueError as exc:
            raise TournamentImportError(
                f"Hoja '{sheet_name}', fila {row_number}: "
                f"fecha inválida '{raw}'. Usa formato YYYY-MM-DD."
            ) from exc

    @staticmethod
    def _normalize_tournament_status(value: str) -> str:
        normalized = value.strip().lower()
        if normalized in {TournamentStatus.DRAFT, TournamentStatus.FINISHED}:
            return normalized
        return TournamentStatus.DRAFT

    @staticmethod
    def _normalize_round_status(value: str) -> str:
        normalized = value.strip().lower()
        if normalized == "draft":
            return RoundStatus.PENDING
        if normalized == RoundStatus.PENDING:
            return RoundStatus.PENDING
        return RoundStatus.FINISHED

    @staticmethod
    def _as_str(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()